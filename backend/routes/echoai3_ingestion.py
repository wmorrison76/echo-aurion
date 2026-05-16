"""
EchoAi3 -- File/PDF Ingestion Pipeline
=========================================
Upload and extract data from PDFs, spreadsheets, and documents.
Supports invoice OCR, menu parsing, financial statement extraction,
vendor catalogs, and BEO document processing.
Integrates with EchoAi3 for intelligent document understanding.
"""
import os
import json
from datetime import datetime, timezone
from uuid import uuid4
from fastapi import APIRouter, UploadFile, File, Form, Query
from typing import Optional

import database
from tamper_audit import log_entry as trace_log

db = database.db
router = APIRouter(prefix="/api/echoai3/ingest", tags=["echoai3-ingestion"])

LLM_KEY = os.environ.get("EMERGENT_LLM_KEY", "")
UPLOAD_DIR = "/app/backend/uploads/ingestion"
os.makedirs(UPLOAD_DIR, exist_ok=True)

_now = lambda: datetime.now(timezone.utc).isoformat()
_uid = lambda: uuid4().hex[:12]

SUPPORTED_TYPES = {
    "application/pdf": "pdf",
    "text/csv": "csv",
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": "xlsx",
    "application/vnd.ms-excel": "xls",
    "text/plain": "txt",
    "application/json": "json",
    "image/png": "png",
    "image/jpeg": "jpeg",
}

DOCUMENT_CATEGORIES = [
    "invoice", "menu", "beo", "financial_statement", "vendor_catalog",
    "recipe", "compliance_report", "employee_document", "contract", "other",
]


def _extract_text_from_pdf(filepath: str) -> str:
    """Extract text from a PDF file."""
    try:
        import fitz  # PyMuPDF
        doc = fitz.open(filepath)
        text = ""
        for page in doc:
            text += page.get_text() + "\n"
        doc.close()
        return text.strip()
    except ImportError:
        # Fallback: try pdfplumber
        try:
            import pdfplumber
            text = ""
            with pdfplumber.open(filepath) as pdf:
                for page in pdf.pages:
                    page_text = page.extract_text()
                    if page_text:
                        text += page_text + "\n"
            return text.strip()
        except ImportError:
            return "[PDF extraction libraries not available. Install PyMuPDF or pdfplumber.]"


def _extract_text_from_csv(filepath: str) -> str:
    """Extract text from CSV."""
    import csv
    rows = []
    with open(filepath, "r", errors="replace") as f:
        reader = csv.reader(f)
        for i, row in enumerate(reader):
            if i > 500:
                break
            rows.append(",".join(row))
    return "\n".join(rows)


def _extract_text_from_txt(filepath: str) -> str:
    """Extract text from plain text file."""
    with open(filepath, "r", errors="replace") as f:
        return f.read()[:50000]


def _extract_text_from_json(filepath: str) -> str:
    """Extract text from JSON file."""
    with open(filepath, "r", errors="replace") as f:
        data = json.load(f)
    return json.dumps(data, indent=2)[:50000]


def _extract_text_from_xlsx(filepath: str) -> str:
    """Extract text from Excel file."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(filepath, read_only=True)
        text_parts = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            text_parts.append(f"--- Sheet: {sheet_name} ---")
            for row in ws.iter_rows(max_row=200, values_only=True):
                vals = [str(c) if c is not None else "" for c in row]
                text_parts.append(",".join(vals))
        wb.close()
        return "\n".join(text_parts)
    except ImportError:
        return "[openpyxl not installed for Excel extraction]"


EXTRACTORS = {
    "pdf": _extract_text_from_pdf,
    "csv": _extract_text_from_csv,
    "txt": _extract_text_from_txt,
    "json": _extract_text_from_json,
    "xlsx": _extract_text_from_xlsx,
    "xls": _extract_text_from_xlsx,
}


async def _ai_classify_document(text: str) -> dict:
    """Use LLM to classify document type and extract structured data."""
    try:
        from emergentintegrations.llm.chat import LlmChat, UserMessage
        chat = LlmChat(
            api_key=LLM_KEY,
            session_id=f"ingest-{_uid()}",
            system_message=(
                "You are EchoAi3's Document Intelligence engine. Analyze the document text and return a JSON object with:\n"
                "1. category: one of [invoice, menu, beo, financial_statement, vendor_catalog, recipe, compliance_report, employee_document, contract, other]\n"
                "2. summary: 2-3 sentence summary of the document\n"
                "3. key_data: extracted key-value pairs (amounts, dates, names, items)\n"
                "4. confidence: 0-100 how confident you are in the classification\n"
                "5. entities: list of named entities found (people, companies, products, amounts)\n"
                "6. actionable_items: list of actions that should be taken based on this document\n"
                "Return ONLY valid JSON, no markdown."
            ),
        )
        chat.with_model("openai", "gpt-4.1-mini")
        response = await chat.send_message(UserMessage(text=f"Classify and extract data from this document:\n\n{text[:8000]}"))

        # Try to parse JSON from response
        try:
            # Handle potential markdown wrapping
            clean = response.strip()
            if clean.startswith("```"):
                clean = clean.split("\n", 1)[1] if "\n" in clean else clean
                clean = clean.rsplit("```", 1)[0] if "```" in clean else clean
            return json.loads(clean)
        except json.JSONDecodeError:
            return {
                "category": "other",
                "summary": response[:300],
                "key_data": {},
                "confidence": 40,
                "entities": [],
                "actionable_items": [],
            }
    except Exception as e:
        return {
            "category": "other",
            "summary": f"AI classification unavailable: {str(e)[:100]}",
            "key_data": {},
            "confidence": 0,
            "entities": [],
            "actionable_items": [],
        }


async def _ai_query_document(text: str, query: str) -> str:
    """Ask a question about a document."""
    try:
        from emergentintegrations.llm.chat import LlmChat, UserMessage
        chat = LlmChat(
            api_key=LLM_KEY,
            session_id=f"doc-query-{_uid()}",
            system_message=(
                "You are EchoAi3's Document Intelligence engine. Answer questions about the document precisely. "
                "Use specific numbers and details from the text. If the answer isn't in the document, say so."
            ),
        )
        chat.with_model("openai", "gpt-4.1-mini")
        return await chat.send_message(UserMessage(text=f"Document:\n{text[:6000]}\n\nQuestion: {query}"))
    except Exception:
        return "AI query unavailable. Document text has been extracted and stored for manual review."


# ─── API Endpoints ───

@router.post("/upload")
async def upload_document(
    file: UploadFile = File(...),
    category: str = Form("auto"),
    tags: str = Form(""),
):
    """Upload and process a document. Extracts text, classifies, and stores."""
    # Validate file type
    content_type = file.content_type or ""
    ext = SUPPORTED_TYPES.get(content_type)
    if not ext:
        # Try from filename
        fname = file.filename or ""
        ext = fname.rsplit(".", 1)[-1].lower() if "." in fname else None
        if ext not in EXTRACTORS:
            return {"error": f"Unsupported file type: {content_type}. Supported: {list(SUPPORTED_TYPES.values())}"}

    doc_id = f"doc-{_uid()}"
    filename = f"{doc_id}.{ext}"
    filepath = os.path.join(UPLOAD_DIR, filename)

    # Save file
    content = await file.read()
    with open(filepath, "wb") as f:
        f.write(content)

    # Extract text
    extractor = EXTRACTORS.get(ext)
    extracted_text = ""
    if extractor:
        extracted_text = extractor(filepath)

    # AI classification
    classification = {}
    if category == "auto" and extracted_text:
        classification = await _ai_classify_document(extracted_text)
        category = classification.get("category", "other")

    # Store metadata
    doc_record = {
        "document_id": doc_id,
        "filename": file.filename,
        "stored_as": filename,
        "content_type": content_type,
        "file_ext": ext,
        "size_bytes": len(content),
        "category": category,
        "tags": [t.strip() for t in tags.split(",") if t.strip()],
        "extracted_text_length": len(extracted_text),
        "classification": classification,
        "summary": classification.get("summary", ""),
        "key_data": classification.get("key_data", {}),
        "entities": classification.get("entities", []),
        "actionable_items": classification.get("actionable_items", []),
        "confidence": classification.get("confidence", 0),
        "status": "processed",
        "uploaded_at": _now(),
    }
    db["ingested_documents"].insert_one(doc_record)
    doc_record.pop("_id", None)

    # Store extracted text separately (can be large)
    db["document_texts"].update_one(
        {"document_id": doc_id},
        {"$set": {"document_id": doc_id, "text": extracted_text, "updated_at": _now()}},
        upsert=True,
    )

    trace_log(
        event_type="document_ingested",
        entity_type="echoai3_ingestion",
        entity_id=doc_id,
        actor_id="echoai3",
        metadata={"filename": file.filename, "category": category, "size": len(content)},
    )

    return {
        "document_id": doc_id,
        "filename": file.filename,
        "category": category,
        "summary": classification.get("summary", "Text extracted, classification pending"),
        "key_data": classification.get("key_data", {}),
        "entities": classification.get("entities", []),
        "actionable_items": classification.get("actionable_items", []),
        "confidence": classification.get("confidence", 0),
        "text_length": len(extracted_text),
        "status": "processed",
    }


@router.get("/documents")
async def list_documents(
    category: str = Query("", description="Filter by category"),
    limit: int = Query(20, ge=1, le=100),
):
    """List ingested documents."""
    query = {}
    if category:
        query["category"] = category

    docs = list(db["ingested_documents"].find(query, {"_id": 0}).sort("uploaded_at", -1).limit(limit))
    return {"documents": docs, "count": len(docs), "categories": DOCUMENT_CATEGORIES}


@router.get("/document/{doc_id}")
async def get_document(doc_id: str, include_text: bool = Query(False)):
    """Get a specific ingested document's metadata and optionally its text."""
    doc = db["ingested_documents"].find_one({"document_id": doc_id}, {"_id": 0})
    if not doc:
        return {"error": "Document not found"}

    if include_text:
        text_doc = db["document_texts"].find_one({"document_id": doc_id}, {"_id": 0})
        doc["extracted_text"] = text_doc.get("text", "") if text_doc else ""

    return doc


@router.post("/document/{doc_id}/query")
async def query_document(doc_id: str, query: str = Form(...)):
    """Ask a question about a specific document."""
    text_doc = db["document_texts"].find_one({"document_id": doc_id}, {"_id": 0})
    if not text_doc or not text_doc.get("text"):
        return {"error": "Document text not found"}

    answer = await _ai_query_document(text_doc["text"], query)
    return {"document_id": doc_id, "query": query, "answer": answer, "timestamp": _now()}


@router.delete("/document/{doc_id}")
async def delete_document(doc_id: str):
    """Delete an ingested document."""
    doc = db["ingested_documents"].find_one({"document_id": doc_id}, {"_id": 0})
    if not doc:
        return {"error": "Document not found"}

    # Delete file
    filepath = os.path.join(UPLOAD_DIR, doc.get("stored_as", ""))
    if os.path.exists(filepath):
        os.remove(filepath)

    db["ingested_documents"].delete_one({"document_id": doc_id})
    db["document_texts"].delete_one({"document_id": doc_id})

    return {"deleted": True, "document_id": doc_id}


@router.get("/stats")
async def ingestion_stats():
    """Get ingestion pipeline statistics."""
    total = db["ingested_documents"].count_documents({})

    # Category breakdown
    pipeline = [
        {"$group": {"_id": "$category", "count": {"$sum": 1}, "total_size": {"$sum": "$size_bytes"}}},
        {"$sort": {"count": -1}},
    ]
    categories = list(db["ingested_documents"].aggregate(pipeline))

    return {
        "total_documents": total,
        "categories": [{
            "category": c["_id"],
            "count": c["count"],
            "total_size_kb": round(c["total_size"] / 1024, 1),
        } for c in categories],
        "supported_formats": list(set(SUPPORTED_TYPES.values())),
        "timestamp": _now(),
    }
